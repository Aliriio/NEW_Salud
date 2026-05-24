# -*- coding: utf-8 -*-
"""Exporta datosProPai de data.js a Excel."""

import json
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE = Path(__file__).resolve().parent
DATA_JS = BASE / "data.js"
OUTPUT = BASE / "plan_cuidados_enfermeria.xlsx"


def cargar_datos():
    texto = DATA_JS.read_text(encoding="utf-8")
    texto = re.sub(r"^export const datosProPai\s*=\s*", "", texto.strip())
    texto = re.sub(r";\s*$", "", texto)
    return json.loads(texto)


def bullets(items):
    if not items:
        return ""
    return "\n".join(f"• {x}" for x in items)


def formato_b6(b6_por_noc):
    if not b6_por_noc:
        return ""
    bloques = []
    for noc, niveles in b6_por_noc.items():
        lineas = [f"[{noc}]"] + [f"  {n}" for n in niveles]
        bloques.append("\n".join(lineas))
    return "\n".join(bloques)


def estilo_encabezado(ws, fila, num_cols, relleno="366092"):
    fill = PatternFill("solid", fgColor=relleno)
    font = Font(bold=True, color="FFFFFF")
    for col in range(1, num_cols + 1):
        cell = ws.cell(fila, col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def ajustar_columnas(ws, anchos):
    for col, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(col)].width = ancho


def hoja_resumen(wb, datos):
    ws = wb.create_sheet("Resumen General", 0)
    ws.merge_cells("A1:D1")
    ws["A1"] = "PLAN DE CUIDADOS DE ENFERMERÍA - RESUMEN GENERAL"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A2:D2")
    ws["A2"] = (
        "Clasificación por área clínica, diagnósticos NANDA, "
        "resultados NOC e intervenciones NIC"
    )
    headers = ["Área Clínica", "Diagnóstico de Enfermería (NANDA)", "Resultados (NOC)", "Intervenciones (NIC)"]
    for col, h in enumerate(headers, 1):
        ws.cell(3, col, h)
    estilo_encabezado(ws, 3, 4)

    fila = 4
    for area, diagnosticos in datos.items():
        primera = True
        for diag, info in diagnosticos.items():
            ws.cell(fila, 1, area if primera else None)
            ws.cell(fila, 2, diag)
            ws.cell(fila, 3, bullets(info.get("noc", [])))
            ws.cell(fila, 4, bullets(info.get("nic", [])))
            for col in range(1, 5):
                ws.cell(fila, col).alignment = Alignment(wrap_text=True, vertical="top")
            primera = False
            fila += 1
    ajustar_columnas(ws, [22, 48, 40, 40])


def hoja_diagnosticos_completos(wb, datos):
    ws = wb.create_sheet("Diagnósticos Completos")
    ws.merge_cells("A1:F1")
    ws["A1"] = "DIAGNÓSTICOS DE ENFERMERÍA - DETALLE COMPLETO"
    ws["A1"].font = Font(bold=True, size=14)
    headers = [
        "Área Clínica",
        "Diagnóstico NANDA",
        "NOC (Resultado)",
        "NIC (Intervención)",
        "Intervenciones Transversales",
        "Escala B6 por NOC",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(2, col, h)
    estilo_encabezado(ws, 2, 6)

    fila = 3
    for area, diagnosticos in datos.items():
        for diag, info in diagnosticos.items():
            ws.cell(fila, 1, area)
            ws.cell(fila, 2, diag)
            ws.cell(fila, 3, bullets(info.get("noc", [])))
            ws.cell(fila, 4, bullets(info.get("nic", [])))
            ws.cell(fila, 5, bullets(info.get("trans", [])))
            ws.cell(fila, 6, formato_b6(info.get("b6_por_noc", {})))
            for col in range(1, 7):
                ws.cell(fila, col).alignment = Alignment(wrap_text=True, vertical="top")
            fila += 1
    ajustar_columnas(ws, [22, 48, 36, 36, 36, 44])


def hoja_escalas_b6(wb, datos):
    ws = wb.create_sheet("Escalas NOC-B6")
    ws.merge_cells("A1:E1")
    ws["A1"] = "ESCALAS DE VALORACIÓN NOC / INDICADOR B6"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A2:E2")
    ws["A2"] = "Puntuación del indicador: 1 (peor) → 5 (óptimo)"
    headers = [
        "Área Clínica",
        "Diagnóstico NANDA",
        "NOC (Resultado Esperado)",
        "Puntuación",
        "Descripción del Nivel",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(3, col, h)
    estilo_encabezado(ws, 3, 5)

    fila = 4
    for area, diagnosticos in datos.items():
        for diag, info in diagnosticos.items():
            b6 = info.get("b6_por_noc") or {}
            if not b6:
                continue
            primera_diag = True
            for noc, niveles in b6.items():
                primera_noc = True
                for nivel in niveles:
                    m = re.match(r"^(\d+)\s*[\.\-]\s*(.+)$", nivel.strip())
                    puntuacion = m.group(1) if m else ""
                    descripcion = nivel.strip()
                    if primera_diag:
                        ws.cell(fila, 1, area)
                        ws.cell(fila, 2, diag)
                        primera_diag = False
                    if primera_noc:
                        ws.cell(fila, 3, noc)
                        primera_noc = False
                    ws.cell(fila, 4, puntuacion)
                    ws.cell(fila, 5, descripcion)
                    for col in range(1, 6):
                        ws.cell(fila, col).alignment = Alignment(wrap_text=True, vertical="top")
                    fila += 1
    ajustar_columnas(ws, [22, 48, 40, 12, 36])


def hoja_transversales(wb, datos):
    ws = wb.create_sheet("Intervenciones Transversales")
    ws.merge_cells("A1:C1")
    ws["A1"] = "INTERVENCIONES TRANSVERSALES POR DIAGNÓSTICO"
    ws["A1"].font = Font(bold=True, size=14)
    headers = ["Área Clínica", "Diagnóstico NANDA", "Intervenciones Transversales"]
    for col, h in enumerate(headers, 1):
        ws.cell(2, col, h)
    estilo_encabezado(ws, 2, 3)

    fila = 3
    for area, diagnosticos in datos.items():
        for diag, info in diagnosticos.items():
            trans = info.get("trans", [])
            if not trans:
                continue
            ws.cell(fila, 1, area)
            ws.cell(fila, 2, diag)
            ws.cell(fila, 3, bullets(trans))
            for col in range(1, 4):
                ws.cell(fila, col).alignment = Alignment(wrap_text=True, vertical="top")
            fila += 1
    ajustar_columnas(ws, [22, 48, 56])


def nombre_hoja(titulo):
    invalidos = r'[]:*?/\\'
    limpio = "".join("-" if c in invalidos else c for c in titulo)
    return limpio[:31] or "Hoja"


def hoja_por_area(wb, area, diagnosticos):
    nombre = nombre_hoja(area)
    ws = wb.create_sheet(nombre)
    ws.merge_cells("A1:D1")
    ws["A1"] = f"ÁREA CLÍNICA: {area.upper()}"
    ws["A1"].font = Font(bold=True, size=14)
    headers = ["Diagnóstico NANDA", "NOC", "NIC", "Intervenciones Transversales"]
    for col, h in enumerate(headers, 1):
        ws.cell(2, col, h)
    estilo_encabezado(ws, 2, 4)

    fila = 3
    for diag, info in diagnosticos.items():
        ws.cell(fila, 1, diag)
        ws.cell(fila, 2, bullets(info.get("noc", [])))
        ws.cell(fila, 3, bullets(info.get("nic", [])))
        ws.cell(fila, 4, bullets(info.get("trans", [])))
        for col in range(1, 5):
            ws.cell(fila, col).alignment = Alignment(wrap_text=True, vertical="top")
        fila += 1
    ajustar_columnas(ws, [48, 40, 40, 44])


def main():
    datos = cargar_datos()
    wb = Workbook()
    wb.remove(wb.active)

    hoja_resumen(wb, datos)
    hoja_diagnosticos_completos(wb, datos)
    hoja_escalas_b6(wb, datos)
    hoja_transversales(wb, datos)
    for area, diagnosticos in datos.items():
        hoja_por_area(wb, area, diagnosticos)

    wb.save(OUTPUT)
    total_diag = sum(len(d) for d in datos.values())
    print(f"Excel generado: {OUTPUT}")
    print(f"  Áreas clínicas: {len(datos)}")
    print(f"  Diagnósticos: {total_diag}")
    print(f"  Hojas: {len(wb.sheetnames)}")


if __name__ == "__main__":
    main()
